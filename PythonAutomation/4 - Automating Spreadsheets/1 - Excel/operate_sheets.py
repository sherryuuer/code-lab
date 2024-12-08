import openpyxl

wb = openpyxl.load_workbook('Automating Worksheets.xlsx')

# 取得有效的 sheets 列表
active_ws = wb.active
print(active_ws.title)

ws = wb['Sheet3']
# 重命名 sheet 名
ws.title = 'Third Sheet'

wb.save('Automating Worksheets.xlsx')

# load
wb = openpyxl.load_workbook('Automating Worksheets.xlsx')
# create sheet
wb.create_sheet('new_sheet')
# delete sheet (delete a list element)
del wb['Third Sheet']
# save
wb.save('Automating Worksheets.xlsx')
