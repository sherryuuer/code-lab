import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.load_workbook('Formulas, Fonts, and Fills.xlsx')

ws = wb['Ratings']

rng = ws['B2:E11']

bad_rating_font = Font(italic=True, color='FF0000', size=12)
good_rating_font = Font(bold=True, color='FFFFFF', size=12)
good_rating_fill = PatternFill(start_color='0000FF', end_color='0000FF', patternType='solid')

for row in rng:
    for cell in row:
        if type(cell.value) == int and cell.value < 5:
            cell.font = bad_rating_font
        elif type(cell.value) == int and cell.value == 10:
            cell.font = good_rating_font
            cell.fill = good_rating_fill


wb.save('Formulas, Fonts, and Fills.xlsx')