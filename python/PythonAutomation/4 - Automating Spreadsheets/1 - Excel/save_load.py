import openpyxl

# create a object and save
wb = openpyxl.Workbook()
wb.save('my_workbook.xlsx')

# open and edit
wb = openpyxl.load_workbook('my_workbook.xlsx')
ws = wb.active
ws['A1'] = 'Hello Excel!'

# save
wb.save('my_workbook.xlsx')
