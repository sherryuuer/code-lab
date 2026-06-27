import openpyxl

wb = openpyxl.load_workbook('Employee Ratings.xlsx')

ws = wb['Ratings']

rng = ws['B2:E11']  # 结果是一个二元嵌套元祖

for row in rng:
    for cell in row:
        val = cell.value

        if type(val) != int:
            cell.value = ''
        elif val < 0:
            cell.value = 0
        elif val > 10:
            cell.value = 10

wb.save('Employee Ratings.xlsx')


# write data to Excel
data = [
    ["Product", "Category", "Price"],
    ["Apple iPhone 12", "Electronics", 799],
    ["Nike Air Force 1", "Footwear", 90],
    ["The Alchemist", "Books", 16.99],
    ["Instant Pot Duo", "Home & Kitchen", 89]
]

wb = openpyxl.load_workbook('Python 2 Excel.xlsx')

ws = wb['Sheet1']

rng_addr = 'A1:C' + str(len(data))  # 取得动态的长度

rng = ws[rng_addr]  # 二元嵌套元祖，元祖是因为，表的位置信息和性质是不变的

for i in range(len(data)):
    row = data[i]

    for j in range(len(row)):
        val = row[j]

        rng[i][j].value = val

wb.save('Python 2 Excel.xlsx')
