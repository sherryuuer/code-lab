import re
import openpyxl

wb = openpyxl.load_workbook('Employees.xlsx')

names = ['John Smith', 'Amit Patel', 'Robert Brown', 'Sanaa Al Farsi', 'Michael Miller',
         'Chiamaka Mbah', 'James Wilson', 'Maria Rodriguez', 'David Clark', 'Chen Wei']

for name in names:
    wb.create_sheet(name)

# check sheets name and delete
wb = openpyxl.load_workbook('Employees.xlsx')
pattern = r'^Sheet\d+$'
ws_names = wb.sheetnames
for ws_name in ws_names:
    if re.search(pattern, ws_name) is not None:
        del wb[ws_name]

wb.save('Employees.xlsx')
