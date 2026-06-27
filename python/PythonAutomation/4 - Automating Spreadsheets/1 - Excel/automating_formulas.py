import openpyxl

wb = openpyxl.load_workbook('Formulas, Fonts, and Fills.xlsx')

ws = wb['Ratings']

rng = ws['B2:E11']

ws['F1'] = 'Average Rating'

for idx, row in enumerate(rng):
    current_row = str(idx + 2)

    formula_cell = 'F' + current_row

    formula = '=AVERAGE(B' + current_row + ':E' + current_row + ')'

    ws[formula_cell] = formula

    for cell in row:
        val = cell.value

        if type(val) != int:
            cell.value = ''
        elif val < 0:
            cell.value = 0
        elif val > 10:
            cell.value = 10


wb.save('Formulas, Fonts, and Fills.xlsx')
