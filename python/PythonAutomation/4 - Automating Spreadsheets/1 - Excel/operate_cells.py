import openpyxl

wb = openpyxl.load_workbook('my_workbook.xlsx')
ws = wb['Sheet1']

c = ws['A1']
c_val = c.value
print(c_val)

reversed = c_val[::-1]
ws['A2'] = reversed

ws.cell(row=11, column=11, value='Greetings human!')
c_val2 = ws.cell(row=11, column=11).value
print(c_val2)

wb.save('my_workbook.xlsx')
